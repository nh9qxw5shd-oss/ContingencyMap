#!/usr/bin/env python3
"""Convert the two ITSR MML contingency-plan workbooks into cmap_plans rows,
following the conventions of the existing MML-0 example in the system:
  - plan_code from the sheet tab (normalised, e.g. "MML -1" -> "MML-1")
  - scenario_group from the banner ("Full Block Plan" -> "Full block")
  - one step per service group: {group, od, action, plan}
  - CANCELLED -> Suspend, AS BOOKED -> Normal working, everything else -> Alteration
  - TLK headcodes get an "xx" suffix and the TLK tag is dropped ("9R TLK" -> "9Rxx"),
    peak markers kept as "(PEAK)"; EMR groups keep their codes with "(EMR)"
  - NORTH/SOUTH-of-block variants are merged into one step with
    "NORTH OF BLOCK: ... / SOUTH OF BLOCK: ..." in the plan text
  - bullet characters stripped; footnote bullet lists -> constraints
"""
import openpyxl, json, re, sys

UP = "/root/.claude/uploads/844e0347-fcc9-50b6-b958-0f65562befcc"
LONDON = f"{UP}/fb79b79b-ITSR__MML_Contingency_Plans_London__BedfordV4.xlsx"
NORTH = f"{UP}/1a9fc283-March_2026_V2__ITSR_MML_North_Contingency_Plans_Bedford__Chesterfield_.xlsx"

# ---------- helpers ----------

def clean(s):
    if s is None:
        return ""
    s = str(s).replace(" ", " ").replace("\t", " ")
    s = re.sub(r"[ ]+", " ", s)
    # strip bullet prefixes at start of each line
    lines = []
    for ln in s.split("\n"):
        ln = re.sub(r"^\s*[·•\-]\s*", "", ln.strip())
        if ln:
            lines.append(ln)
    return "\n".join(lines).strip()

def oneline(s, sep="; "):
    return sep.join(x.strip() for x in clean(s).split("\n") if x.strip())

def summary_join(s):
    """Join wrapped restriction lines: continue a phrase where the wrap is
    mid-sentence, otherwise separate distinct notes with ', '."""
    lines = [l.strip() for l in clean(s).split("\n") if l.strip()]
    out = []
    for ln in lines:
        if out and (out[-1].rstrip().lower().endswith((" to", " and", " or", " of", " the", "-", "–", "/"))
                    or ln[0].islower()):
            out[-1] = out[-1] + " " + ln
        else:
            out.append(ln)
    return ", ".join(out)

def flat(s):
    return re.sub(r"\s+", " ", clean(s).replace("\n", " ")).strip()

def is_bullet_row(raw):
    return raw is not None and str(raw).lstrip().startswith(("\u2022", "•"))

class Sheet:
    """Merged-cell-aware accessor."""
    def __init__(self, ws):
        self.ws = ws
        self.master = {}   # (row,col) -> (mrow,mcol) top-left of merge
        self.span = {}     # (mrow,mcol) -> (min_row,max_row)
        for rng in ws.merged_cells.ranges:
            for r in range(rng.min_row, rng.max_row + 1):
                for c in range(rng.min_col, rng.max_col + 1):
                    self.master[(r, c)] = (rng.min_row, rng.min_col)
            self.span[(rng.min_row, rng.min_col)] = rng

    def val(self, row, col):
        m = self.master.get((row, col), (row, col))
        return self.ws.cell(m[0], m[1]).value

    def starts_here(self, row, col):
        """True if (row,col) is the top-left of its merge (or unmerged)."""
        m = self.master.get((row, col))
        return m is None or m == (row, col)

    def raw(self, row, col):
        return self.ws.cell(row, col).value

COL = {c: i + 1 for i, c in enumerate("ABCDEFGHIJKL")}

# ---------- action + status ----------

CANON_STATUS = {
    "cancelled", "amended", "split", "split service", "as booked",
    "run as booked", "runs as booked", "as-booked",
}
PAREN_ONLY = re.compile(r"\(([^)]*)\)")

def action_for(status, plan):
    s = oneline(status).lower()
    core = PAREN_ONLY.sub("", s).strip()
    if not s:
        return "Normal working" if not plan else "Normal working"
    if "cancel" in core and len(core) <= len("cancelled") + 2:
        return "Suspend"
    if core in ("as booked", "run as booked", "runs as booked", "normal service"):
        return "Normal working"
    return "Alteration"

def status_is_informative(status):
    """Does the status text carry detail beyond the action keyword?"""
    s = oneline(status)
    core = PAREN_ONLY.sub("", s).strip().lower().rstrip(".")
    return bool(core) and core not in CANON_STATUS

def step_plan(status, plan):
    """Compose plan text: keep informative status detail, then operational plan."""
    parts = []
    if status_is_informative(status):
        parts.append(oneline(status, sep=" — ") if len(clean(status).split("\n")) > 1 else clean(status))
    p = clean(plan)
    if p:
        parts.append(p)
    return "\n\n".join(parts)

# ---------- group / od normalisation ----------

HEADCODE = re.compile(r"^\d[A-Z](?:xx)?(?:/\d?[A-Z](?:xx)?)*$", re.I)

def norm_group_london(text):
    """'9R TLK' -> '9Rxx'; '9K TLK (PEAK)' -> '9Kxx (PEAK)'; '1C/1F EMR' -> '1C/1F (EMR)'."""
    t = clean(text).replace("\n", " ")
    t = re.sub(r"[ ]+", " ", t).strip()
    if not t:
        return ""
    if re.search(r"\bEMR\b", t):
        rest = re.sub(r"\(?\bEMR\b\)?", "", t).strip()
        rest = re.sub(r"[ ]+", " ", rest)
        if not rest:
            return "EMR"
        if rest.upper().startswith("ALL "):
            return t  # e.g. "ALL EMR SERVICES"
        return f"{rest} (EMR)"
    peak = ""
    m = re.search(r"\(([^)]*PEAK[^)]*)\)", t, re.I)
    if m:
        peak = f" ({m.group(1).strip().upper()})"
        t = t.replace(m.group(0), "")
    t = re.sub(r"\bTLK\b", "", t).strip()
    t = re.sub(r"[ ]+", " ", t)
    if HEADCODE.match(t) and not t.lower().endswith("xx"):
        t = t + "xx"
    return (t + peak).strip()

def norm_group_north(text):
    t = clean(text)
    return " / ".join(x.strip() for x in t.split("\n") if x.strip())

def norm_od_north(text):
    lines = [x.strip() for x in clean(text).split("\n") if x.strip()]
    if not lines:
        return ""
    if len(lines) == 1:
        return lines[0]
    return f"{lines[0]} ({' / '.join(lines[1:])})"

# ---------- titles ----------

SMALL = {"to", "and", "or", "via", "of", "the", "at", "in", "from", "for", "with"}
FIX = {"st": "St", "jn": "Jn", "jct": "Jct", "emr": "EMR", "tlk": "TLK", "xc": "XC",
       "na": "N/A", "itsr": "ITSR", "tph": "TPH", "hl": "HL", "ll": "LL"}

def titlecase(s):
    out = []
    for i, w in enumerate(re.split(r"(\s+|\(|\)|/|–|-)", s)):
        lw = w.lower()
        if lw in FIX:
            out.append(FIX[lw])
        elif i > 0 and lw in SMALL:
            out.append(lw)
        elif w and w[0].isalpha():
            out.append(w[0].upper() + w[1:].lower())
        else:
            out.append(w)
    return "".join(out)

def title_from_banner(b3):
    t = clean(b3).replace("\n", " ")
    # drop the "Midland Mainline North 4a –" / "Midland Mainline North 13a -" prefix
    t = re.sub(r"^\s*Midland Mainline(\s+North)?\s+N?\d+[a-cA-C]?\s*[–—-]?\s*", "", t, flags=re.I)
    t = re.sub(r"\s+", " ", t).strip()
    return titlecase(t)

def scenario_from(b4):
    s = clean(b4).lower()
    if "full" in s:
        return "Full block"
    if "partial" in s:
        return "Partial block"
    return "Degraded conditions"

# ---------- London full-block style sheets ----------

def parse_london_std(sh, plan_col="G", last_row=None):
    """Sheets with C=group, D=od, E=status(+overflow rows), G(=or G:H)=plan."""
    ws = sh.ws
    pc = COL[plan_col]
    rows = range(6, (last_row or ws.max_row) + 1)
    anchors = []
    for r in rows:
        dv = sh.raw(r, COL["D"])
        if dv is not None and clean(dv) and sh.starts_here(r, COL["D"]):
            anchors.append(r)
        else:
            # block with no O/D (e.g. MML-8 9T row): status starts, C has a headcode
            ev = sh.raw(r, COL["E"])
            cv = clean(sh.raw(r, COL["C"]) or "")
            if (ev is not None and clean(ev) and sh.starts_here(r, COL["E"])
                    and sh.val(r, COL["D"]) in (None, "") and cv and cv not in ("TLK", "EMR")
                    and not cv.startswith("(")):
                anchors.append(r)
    steps = []
    for i, a in enumerate(anchors):
        end = (anchors[i + 1] - 1) if i + 1 < len(anchors) else (last_row or ws.max_row)
        group, status, plans = [], [], []
        for r in range(a, end + 1):
            for colname, bucket in (("C", group), ("E", status), (plan_col, plans)):
                v = sh.raw(r, COL[colname])
                if v is not None and clean(v) and sh.starts_here(r, COL[colname]):
                    bucket.append(flat(v))
        od = clean(sh.val(a, COL["D"]) or "")
        st = " ".join(status)
        steps.append({
            "group": norm_group_london(" ".join(group)),
            "od": od.replace("\n", " "),
            "action": action_for(st, "\n".join(plans)),
            "plan": step_plan(st, "\n".join(plans)),
        })
    return steps

# ---------- London stacked partial sheets (MML-4a / 6a / 7a) ----------

def parse_london_stacked(sh, plan_col="E"):
    ws = sh.ws
    # locate the two header rows (C = "Service Group")
    headers = [r for r in range(1, ws.max_row + 1)
               if clean(sh.val(r, COL["C"])) == "Service Group"]
    tables, footnotes = [], []
    for h in headers:
        rows = []
        r = h + 1
        while r <= ws.max_row:
            c = clean(sh.val(r, COL["C"]) or "")
            d = clean(sh.val(r, COL["D"]) or "")
            if not c and not d:
                break
            if c.startswith("***") or c.startswith("•"):
                footnotes.append(c)
                break
            if c == "Service Group":
                break
            if sh.starts_here(r, COL["C"]) and c:
                rows.append((c, d, clean(sh.val(r, COL[plan_col]) or "")))
            r += 1
        label = ""
        restr = clean(sh.val(h + 1, COL["B"]) or "").lower()
        if "fast" in restr:
            label = "FAST LINES BLOCKED"
        elif "slow" in restr:
            label = "SLOW LINES BLOCKED"
        tables.append((label, rows, clean(sh.val(h + 1, COL["B"]) or "")))
    # merge by (group, od)
    order, bykey = [], {}
    for label, rows, _restr in tables:
        for c, d, plan in rows:
            key = (norm_group_london(c), d.replace("\n", " "))
            if key not in bykey:
                bykey[key] = []
                order.append(key)
            bykey[key].append((label, plan))
    steps = []
    for key in order:
        entries = bykey[key]
        plans = [p for _, p in entries]
        low = [p.lower().rstrip(". ") for p in plans]
        if all(p.startswith("run as booked") or p == "as booked" for p in low):
            action, plan = "Normal working", ""
            extra = {p for p in plans if p.lower().rstrip(". ") not in ("run as booked", "as booked")}
            if extra:
                plan = " / ".join(sorted(extra))
        elif all(p == "cancelled" for p in low):
            action, plan = "Suspend", ""
        elif len(set(low)) == 1:
            action = action_for(plans[0], "")
            plan = plans[0] if plans[0].lower() not in ("cancelled",) else ""
        else:
            action = "Alteration"
            plan = "\n\n".join(f"{lbl}: {p}" if lbl else p for lbl, p in entries)
        steps.append({"group": key[0], "od": key[1], "action": action, "plan": plan})
    restrictions = " / ".join(dict.fromkeys(summary_join(t[2]) for t in tables))
    return steps, footnotes, restrictions

# ---------- MML-2a (colour-coded) ----------

def parse_mml2a(ws):
    sh = Sheet(ws)
    steps, order, bykey = [], [], {}
    for block in (range(6, 17), range(22, 33)):
        for r in block:
            c = clean(sh.val(r, COL["C"]) or "")
            d = clean(sh.val(r, COL["D"]) or "")
            if not c:
                continue
            note = ""
            amended = False
            for col in ("F", "G"):
                cell = ws.cell(r, COL[col])
                rgb = cell.fill.fgColor.rgb if cell.fill and cell.fill.fgColor else ""
                if isinstance(rgb, str) and rgb.endswith("FFFF00"):
                    amended = True
                v = clean(cell.value or "")
                if v:
                    note = v
            key = (norm_group_london(c), d.replace("\n", " "))
            if key not in bykey:
                bykey[key] = (amended, note)
                order.append(key)
            else:
                pa, pn = bykey[key]
                bykey[key] = (pa or amended, pn or note)
    for key in order:
        amended, note = bykey[key]
        steps.append({
            "group": key[0], "od": key[1],
            "action": "Alteration" if amended else "Normal working",
            "plan": note,
        })
    return steps

# ---------- MML-7b ----------

def parse_mml7b(ws):
    sh = Sheet(ws)
    steps = []
    for r in range(6, 17):
        c = clean(sh.val(r, COL["C"]) or "")
        if not c or not sh.starts_here(r, COL["C"]):
            continue
        d = clean(sh.val(r, COL["D"]) or "")
        f = clean(sh.val(r, COL["F"]) or "")
        g = clean(sh.val(r, COL["G"]) or "")
        steps.append({
            "group": norm_group_london(c),
            "od": d.replace("\n", " "),
            "action": "Alteration" if f else "Normal working",
            "plan": step_plan(f, g) if f else g,
        })
    return steps

# ---------- North sheets ----------

def north_layout(sh):
    """wide: NORTH/SOUTH in col C with D=group; narrow: C=group."""
    return "wide" if clean(sh.val(5, COL["D"])) == "Service Group" else "narrow"

def parse_north(ws):
    sh = Sheet(ws)
    layout = north_layout(sh)
    if layout == "wide":
        gcol, odcol, stcol, plcol = "D", "E", "F", "H"
    else:
        gcol, odcol, stcol, plcol = "C", "D", "E", "G"

    constraints, notes = [], []
    entries = []  # (side, group, od, status, plan)
    side = ""
    last_data_row = ws.max_row
    # the final considerations row: group-col cell merged across to plan col with bullets
    for r in range(6, ws.max_row + 1):
        raw = sh.raw(r, COL[gcol])
        if is_bullet_row(raw) and sh.starts_here(r, COL[gcol]):
            constraints.append(clean(raw))
            last_data_row = min(last_data_row, r - 1)
    for r in range(6, last_data_row + 1):
        if layout == "wide":
            cv = clean(sh.raw(r, COL["C"]) or "")
            if cv and sh.starts_here(r, COL["C"]):
                if cv.upper() in ("NORTH", "SOUTH"):
                    side = cv.upper()
                else:
                    notes.append(cv)  # e.g. "BLOCK between Bedford and Kettering"
        g = sh.raw(r, COL[gcol])
        if g is None or not clean(g) or not sh.starts_here(r, COL[gcol]):
            continue
        if is_bullet_row(g):
            continue
        od = clean(sh.val(r, COL[odcol]) or "")
        st = flat(sh.val(r, COL[stcol]) or "")
        pl = flat(sh.val(r, COL[plcol]) or "")
        entries.append((side, norm_group_north(g), norm_od_north(od), st, pl))

    # merge NORTH/SOUTH duplicates
    order, bykey = [], {}
    for side, g, od, st, pl in entries:
        key = (g, od)
        if key not in bykey:
            bykey[key] = []
            order.append(key)
        bykey[key].append((side, st, pl))
    steps = []
    for key in order:
        items = bykey[key]
        if len(items) == 1:
            side, st, pl = items[0]
            steps.append({"group": key[0], "od": key[1],
                          "action": action_for(st, pl), "plan": step_plan(st, pl)})
        else:
            actions = {action_for(st, pl) for _, st, pl in items}
            action = actions.pop() if len(actions) == 1 else "Alteration"
            parts = []
            for side, st, pl in items:
                bits = []
                if status_is_informative(st):
                    bits.append(flat(st))
                elif not pl:
                    bits.append(flat(st))
                if pl:
                    bits.append(pl)
                detail = " \u2014 ".join(b for b in bits if b) or "As booked"
                lbl = f"{side} OF BLOCK: " if side else ""
                parts.append(lbl + detail)
            steps.append({"group": key[0], "od": key[1], "action": action,
                          "plan": "\n\n".join(parts)})
    return steps, constraints, notes

# ---------- drive ----------

def plan_shell(code, ws):
    sh = Sheet(ws)
    b3 = sh.val(3, COL["B"])
    b4 = sh.val(4, COL["B"])
    return {
        "plan_code": code,
        "title": title_from_banner(b3),
        "scenario_group": scenario_from(b4),
        "summary": "", "assumptions": "", "constraints": "",
        "steps": [], "docs": [],
    }

def restriction(ws, col="B", row=6):
    sh = Sheet(ws)
    return summary_join(sh.val(row, COL[col]) or "")

plans = []

wbL = openpyxl.load_workbook(LONDON, data_only=True)
wbLf = openpyxl.load_workbook(LONDON)  # with styles for MML-2a colours

london_std = {
    "MML -1": ("MML-1", "G", None),
    "MML -2": ("MML-2", "G", None),
    "MML-3": ("MML-3", "G", None),
    "MML -4": ("MML-4", "G", None),
    "MML -5": ("MML-5", "G", None),
    "MML -5a": ("MML-5a", "G", None),
    "MML -6": ("MML-6", "G", None),
    "MML -7": ("MML-7", "G", None),
    "MML -8": ("MML-8", "G", None),
    "MML -9": ("MML-9", "G", None),
}
for tab, (code, plan_col, last) in london_std.items():
    ws = wbL[tab]
    p = plan_shell(code, ws)
    p["summary"] = restriction(ws)
    p["steps"] = parse_london_std(Sheet(ws), plan_col, last)
    plans.append(p)

# MML-2a (colour-coded)
ws = wbLf["MML-2a"]
p = plan_shell("MML-2a", ws)
p["summary"] = ("Bedford to Luton reduced to a 2 or 3 track railway (fast or slow lines blocked). "
                "All services run as booked except where noted.")
p["steps"] = parse_mml2a(ws)
plans.append(p)

# stacked text tables
for tab, code in (("MML -4a", "MML-4a"), ("MML -6a", "MML-6a"), ("MML -7a", "MML-7a")):
    ws = wbL[tab]
    p = plan_shell(code, ws)
    steps, foot, restr = parse_london_stacked(Sheet(ws))
    p["steps"] = steps
    p["summary"] = restr
    p["constraints"] = "\n".join(dict.fromkeys(foot))
    plans.append(p)

# MML-7b
ws = wbL["MML -7b"]
p = plan_shell("MML-7b", ws)
p["summary"] = restriction(ws)
p["steps"] = parse_mml7b(ws)
plans.append(p)

# sort London plans into natural order
def plan_key(p):
    m = re.match(r"MML-(N?)(\d+)([a-z]?)", p["plan_code"])
    return (m.group(1), int(m.group(2)), m.group(3))
plans.sort(key=plan_key)

# North workbook
wbN = openpyxl.load_workbook(NORTH, data_only=True)
for tab in wbN.sheetnames:
    if tab == "Index":
        continue
    ws = wbN[tab]
    p = plan_shell(tab, ws)  # tabs already "MML-N1" style
    steps, constraints, notes = parse_north(ws)
    p["steps"] = steps
    p["summary"] = restriction(ws)
    p["constraints"] = "\n".join(constraints)
    p["assumptions"] = "\n".join(notes)
    plans.append(p)

for i, p in enumerate(plans):
    p["sort_order"] = 100 + i * 10

with open("plans.json", "w") as f:
    json.dump(plans, f, indent=1, ensure_ascii=False)

print(f"{len(plans)} plans, {sum(len(p['steps']) for p in plans)} steps")
for p in plans:
    print(f"{p['plan_code']:10s} {p['scenario_group']:14s} steps={len(p['steps']):2d}  {p['title'][:70]}")
